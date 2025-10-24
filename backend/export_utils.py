"""
Export utilities for generating PDF and XLSX reports
"""
from datetime import datetime
from typing import List, Dict, Any
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def format_currency(value: float, locale: str = 'en') -> str:
    """Format currency value"""
    if locale == 'it':
        return f"€{value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    return f"€{value:,.2f}"


def generate_daily_prep_pdf(
    data: List[Dict[str, Any]],
    date: str,
    restaurant_name: str,
    locale: str = 'en'
) -> bytes:
    """Generate Daily Preparations PDF export"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch,
                           topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    # Container for elements
    elements = []
    styles = getSampleStyleSheet()
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#059669'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Header
    title_text = "Daily Preparations" if locale == 'en' else "Preparazioni Giornaliere"
    title = Paragraph(f"<b>{restaurant_name}</b><br/>{title_text}", title_style)
    elements.append(title)
    
    date_text = f"Date: {date}" if locale == 'en' else f"Data: {date}"
    date_para = Paragraph(date_text, styles['Normal'])
    elements.append(date_para)
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data with ALL required columns
    if locale == 'en':
        table_data = [['Date', 'Preparation', 'Forecast', 'Available', 'To Make', 'Unit', 'Shelf Life', 'Cost/Portion', 'Est. Total', 'Notes']]
    else:
        table_data = [['Data', 'Preparazione', 'Previsto', 'Disponibile', 'Da Fare', 'Unità', 'Scadenza', 'Costo/Porz.', 'Tot. Stim.', 'Note']]
    
    total_cost = 0
    total_to_make = 0
    
    for item in data:
        shelf_life = item.get('shelfLife', '-')
        cost_per_portion = item.get('costPerPortion', 0)
        total_item_cost = item.get('totalCost', 0)
        notes = item.get('notes', '-')
        
        table_data.append([
            item.get('date', date),
            item.get('name', ''),
            str(round(item.get('forecast', 0), 1)),
            str(round(item.get('available', 0), 1)),
            str(round(item.get('toMake', 0), 1)),
            item.get('unit', 'portions'),
            shelf_life,
            format_currency(cost_per_portion, locale),
            format_currency(total_item_cost, locale),
            (notes[:15] + '...' if len(notes) > 15 else notes) if notes else '-'
        ])
        
        total_cost += total_item_cost
        total_to_make += item.get('toMake', 0)
    
    # Totals row
    totals_label = 'TOTALS' if locale == 'en' else 'TOTALI'
    table_data.append([totals_label, '', '', '', str(round(total_to_make, 1)), '', '', '', format_currency(total_cost, locale), ''])
    
    # Create table with adjusted column widths
    table = Table(table_data, colWidths=[0.6*inch, 1.3*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.9*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (8, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0fdf4')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9fafb')])
    ]))
    
    elements.append(table)
    
    # Footer with page number
    elements.append(Spacer(1, 0.3*inch))
    footer_text = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    footer = Paragraph(footer_text, styles['Italic'])
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    
    return buffer.getvalue()


def generate_daily_prep_xlsx(
    data: List[Dict[str, Any]],
    date: str,
    restaurant_name: str,
    locale: str = 'en'
) -> bytes:
    """Generate Daily Preparations XLSX export with ALL required columns"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"DailyPrep_{date}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    total_fill = PatternFill(start_color="f0fdf4", end_color="f0fdf4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title and date
    ws.merge_cells('A1:J1')
    ws['A1'] = f"{restaurant_name} - Daily Preparations"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Date: {date}"
    ws['A2'].font = Font(bold=True)
    
    # Headers with ALL required columns
    if locale == 'en':
        headers = ['Date', 'Preparation', 'Forecast', 'Available', 'To Make', 'Unit', 'Shelf Life', 'Cost/Portion', 'Est. Total', 'Notes']
    else:
        headers = ['Data', 'Preparazione', 'Previsto', 'Disponibile', 'Da Fare', 'Unità', 'Scadenza', 'Costo/Porz.', 'Tot. Stim.', 'Note']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    total_cost = 0
    total_to_make = 0
    
    for idx, item in enumerate(data, 5):
        ws.cell(row=idx, column=1, value=item.get('date', date))
        ws.cell(row=idx, column=2, value=item.get('name', ''))
        ws.cell(row=idx, column=3, value=round(item.get('forecast', 0), 1))
        ws.cell(row=idx, column=4, value=round(item.get('available', 0), 1))
        ws.cell(row=idx, column=5, value=round(item.get('toMake', 0), 1))
        ws.cell(row=idx, column=6, value=item.get('unit', 'portions'))
        ws.cell(row=idx, column=7, value=item.get('shelfLife', '-'))
        
        cost_per = item.get('costPerPortion', 0)
        total_item = item.get('totalCost', 0)
        
        ws.cell(row=idx, column=8, value=cost_per)
        ws.cell(row=idx, column=8).number_format = '€#,##0.00'
        
        ws.cell(row=idx, column=9, value=total_item)
        ws.cell(row=idx, column=9).number_format = '€#,##0.00'
        
        ws.cell(row=idx, column=10, value=item.get('notes', ''))
        
        total_cost += total_item
        total_to_make += item.get('toMake', 0)
        
        # Borders
        for col in range(1, 11):
            ws.cell(row=idx, column=col).border = border
    
    # Totals row
    totals_row = len(data) + 5
    totals_label = 'TOTALS' if locale == 'en' else 'TOTALI'
    ws.cell(row=totals_row, column=1, value=totals_label)
    ws.cell(row=totals_row, column=5, value=round(total_to_make, 1))
    ws.cell(row=totals_row, column=9, value=total_cost)
    ws.cell(row=totals_row, column=9).number_format = '€#,##0.00'
    
    for col in range(1, 11):
        cell = ws.cell(row=totals_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 20
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_purchase_orders_pdf(
    data: Dict[str, List[Dict[str, Any]]],  # Grouped by supplier
    date: str,
    restaurant_name: str,
    locale: str = 'en'
) -> bytes:
    """Generate Purchase Orders PDF export (grouped by supplier) with ALL required columns"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.4*inch, leftMargin=0.4*inch,
                           topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#059669'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Header
    title_text = "Purchase Orders" if locale == 'en' else "Ordini di Acquisto"
    title = Paragraph(f"<b>{restaurant_name}</b><br/>{title_text}", title_style)
    elements.append(title)
    
    date_text = f"Date: {date}" if locale == 'en' else f"Data: {date}"
    date_para = Paragraph(date_text, styles['Normal'])
    elements.append(date_para)
    elements.append(Spacer(1, 0.3*inch))
    
    grand_total = 0
    
    # Headers with ALL required columns: Supplier, Item (code), Qty+Unit, Unit price, Extended cost, Delivery date, Reason/driver, Notes
    if locale == 'en':
        headers = ['Supplier', 'Item (Code)', 'Qty', 'Unit', 'Unit Price', 'Extended', 'Delivery', 'Reason', 'Notes']
    else:
        headers = ['Fornitore', 'Articolo (Cod.)', 'Qtà', 'Unità', 'Prezzo', 'Totale', 'Consegna', 'Motivo', 'Note']
    
    # Group by supplier
    for supplier_name, items in data.items():
        # Supplier section header
        supplier_para = Paragraph(f"<b>{'Supplier' if locale == 'en' else 'Fornitore'}: {supplier_name}</b>", styles['Heading3'])
        elements.append(supplier_para)
        elements.append(Spacer(1, 0.1*inch))
        
        # Table data
        table_data = [headers]
        supplier_total = 0
        
        for item in items:
            qty = item.get('qtyToOrder', 0)
            unit_price = item.get('unitPrice', 0)
            extended = item.get('extendedCost', qty * unit_price)
            supplier_total += extended
            
            item_code = item.get('code', '')
            item_display = f"{item.get('itemName', '')} ({item_code})" if item_code else item.get('itemName', '')
            
            table_data.append([
                supplier_name,
                item_display[:25],  # Item with code
                str(round(qty, 1)),
                item.get('unit', ''),
                format_currency(unit_price, locale),
                format_currency(extended, locale),
                item.get('deliveryDate', '-')[:10] if item.get('deliveryDate') else '-',
                item.get('driver', '-')[:18] if item.get('driver') else '-',
                item.get('notes', '-')[:12] if item.get('notes') else '-'
            ])
        
        # Supplier subtotal row
        subtotal_label = 'SUBTOTAL' if locale == 'en' else 'SUBTOTALE'
        table_data.append([subtotal_label, '', '', '', '', format_currency(supplier_total, locale), '', '', ''])
        grand_total += supplier_total
        
        # Create table with adjusted widths for 9 columns
        table = Table(table_data, colWidths=[0.9*inch, 1.4*inch, 0.4*inch, 0.4*inch, 0.7*inch, 0.7*inch, 0.7*inch, 1.0*inch, 0.7*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (5, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fef3c7')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.2*inch))
        
        grand_total += supplier_total
    
    # Grand total
    grand_total_label = 'GRAND TOTAL' if locale == 'en' else 'TOTALE GENERALE'
    grand_total_para = Paragraph(
        f"<b>{grand_total_label}: {format_currency(grand_total, locale)}</b>",
        styles['Heading2']
    )
    elements.append(grand_total_para)
    
    # Footer
    elements.append(Spacer(1, 0.3*inch))
    footer_text = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    footer = Paragraph(footer_text, styles['Italic'])
    elements.append(footer)
    
    doc.build(elements)
    return buffer.getvalue()


def generate_purchase_orders_xlsx(
    data: Dict[str, List[Dict[str, Any]]],
    date: str,
    restaurant_name: str,
    locale: str = 'en'
) -> bytes:
    """Generate Purchase Orders XLSX export"""
    wb = Workbook()
    
    # Main sheet
    ws = wb.active
    ws.title = f"PurchaseOrders_{date}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    subtotal_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    supplier_fill = PatternFill(start_color="e0f2fe", end_color="e0f2fe", fill_type="solid")
    grand_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = f"{restaurant_name} - Purchase Orders"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Date: {date}"
    ws['A2'].font = Font(bold=True)
    
    current_row = 4
    grand_total = 0
    
    # Headers with ALL required columns
    if locale == 'en':
        headers = ['Supplier', 'Item (Code)', 'Qty', 'Unit', 'Unit Price', 'Extended Cost', 'Delivery Date', 'Reason/Driver', 'Notes']
    else:
        headers = ['Fornitore', 'Articolo (Cod.)', 'Qtà', 'Unità', 'Prezzo', 'Costo Totale', 'Data Cons.', 'Motivo', 'Note']
    
    for supplier_name, items in data.items():
        # Supplier section header
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = f"{'Supplier' if locale == 'en' else 'Fornitore'}: {supplier_name}"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        ws[f'A{current_row}'].fill = supplier_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
        current_row += 1
        
        # Column headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        current_row += 1
        
        # Data rows
        supplier_total = 0
        for item in items:
            qty = item.get('qtyToOrder', 0)
            unit_price = item.get('unitPrice', 0)
            extended = item.get('extendedCost', qty * unit_price)
            supplier_total += extended
            
            item_code = item.get('code', '')
            item_display = f"{item.get('itemName', '')} ({item_code})" if item_code else item.get('itemName', '')
            
            ws.cell(row=current_row, column=1, value=supplier_name)
            ws.cell(row=current_row, column=2, value=item_display)
            ws.cell(row=current_row, column=3, value=round(qty, 1))
            ws.cell(row=current_row, column=4, value=item.get('unit', ''))
            ws.cell(row=current_row, column=5, value=unit_price)
            ws.cell(row=current_row, column=5).number_format = '€#,##0.00'
            ws.cell(row=current_row, column=6, value=extended)
            ws.cell(row=current_row, column=6).number_format = '€#,##0.00'
            ws.cell(row=current_row, column=7, value=item.get('deliveryDate', ''))
            ws.cell(row=current_row, column=8, value=item.get('driver', ''))
            ws.cell(row=current_row, column=9, value=item.get('notes', ''))
            
            for col in range(1, 10):
                ws.cell(row=current_row, column=col).border = border
            
            current_row += 1
        
        # Subtotal
        subtotal_label = 'SUBTOTAL' if locale == 'en' else 'SUBTOTALE'
        ws.cell(row=current_row, column=1, value=subtotal_label)
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=6, value=supplier_total)
        ws.cell(row=current_row, column=6).number_format = '€#,##0.00'
        ws.cell(row=current_row, column=6).font = Font(bold=True)
        
        for col in range(1, 10):
            ws.cell(row=current_row, column=col).fill = subtotal_fill
            ws.cell(row=current_row, column=col).border = border
        
        current_row += 2
        grand_total += supplier_total
    
    # Grand total
    ws.merge_cells(f'A{current_row}:E{current_row}')
    grand_total_label = 'GRAND TOTAL' if locale == 'en' else 'TOTALE GENERALE'
    ws[f'A{current_row}'] = grand_total_label
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
    ws.cell(row=current_row, column=6, value=grand_total)
    ws.cell(row=current_row, column=6).number_format = '€#,##0.00'
    ws.cell(row=current_row, column=6).font = Font(bold=True, size=12)
    
    for col in range(1, 10):
        ws.cell(row=current_row, column=col).fill = grand_fill
        ws.cell(row=current_row, column=col).border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 25
    
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
